import openpyxl,sys
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from openpyxl.chart import Reference, Series, BarChart

def multTable(N):
    # this creates a excel workbook with 'Sheet' worksheet
    wb=openpyxl.Workbook()
    fontObj=Font(name='Times New Roman', size=20, bold=True,italic=True)

    sheet=wb[wb.sheetnames[0]]
    for i in range(N):
        sheet[f'{get_column_letter(i+2)}1']=i+1
        sheet[f'A{i+2}']=i+1
    for i in range(N):
        for j in range(N):
            sheet[f'{get_column_letter(i+2)}{j+2}']=sheet[f'A{j+2}'].value*sheet[f'{get_column_letter(i+2)}1'].value
            '''
            styling of excel cells
            '''
            if i==j:
                sheet[f'{get_column_letter(j+2)}{j+2}'].font=fontObj
                '''
                setting row and column height
                '''
                sheet.row_dimensions[j+2].height=40
                sheet.column_dimensions[get_column_letter(j+2)].width=20

    '''
    formulas in excel using openpyxl
    '''
    sheet[f'A{N+2}']='Col Total'
    sheet[f'{get_column_letter(N+2)}1']='Row Total'
    
    letter_2=get_column_letter(N+2)
    for i in range(N):
        letter_1=get_column_letter(i+2)
        # assign column sum
        sheet[f'{letter_1}{N+2}']=f'=SUM({letter_1}2:{letter_1}{N+1})'
        # assign row sum
        sheet[f'{letter_2}{i+2}']=f'=SUM(B{i+2}:{get_column_letter(N+1)}{i+2})'

    '''
    merge or unmerge cells
    '''
    sheet.merge_cells(f'{get_column_letter(N+3)}1:{get_column_letter(N+5)}{N+2}')
    sheet.merge_cells(f'{get_column_letter(1)}{N+3}:{get_column_letter(N+2)}{N+5}')

    '''
    freeze starting panes
    '''
    sheet.freeze_panes='B2'
    '''
    create charts
    link: https://openpyxl.readthedocs.io/en/stable/charts/chart_layout.html#size-and-position
    '''
    referObj=Reference(sheet,min_col=2,min_row=2,max_row=N+1)
    seriesObj=Series(referObj,title='sample barchart')
    chartObj=BarChart()
    chartObj.series.append(seriesObj)
    chartObj.x_axis.title='x-axis'
    chartObj.y_axis.title='y-axis'
    chartObj.style=13
    sheet.add_chart(chartObj,f'{get_column_letter(N+5)}{N+5}')

    wb.save('multiplicationTable.xlsx')

if __name__=="__main__":
    if len(sys.argv)>1:
        N=sys.argv[1]
        if N.isnumeric():
            N=int(N)
        else:
            print('Give valid N.')
            sys.exit()
    else:
        print('Give N.')
        sys.exit()

    multTable(N)