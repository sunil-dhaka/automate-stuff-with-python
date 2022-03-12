import openpyxl,sys

def blankRows(file):
    wb=openpyxl.load_workbook(file)
    sheet=wb[wb.sheetnames[0]]

    new_wb=openpyxl.Workbook()
    new_sheet=new_wb[new_wb.sheetnames[0]]

    '''
    simple re-write approach
    TODO: excel does not allow more than 16384 columns so this throws error if there are more thatn 16384 rows in original file    
    link of issue: https://stackoverflow.com/questions/45545755/openpyxl-write-list-of-numeric-data-value-error-invalid-column-index
    '''
    N=sheet.max_column
    M=sheet.max_row
    print(M,N)
    for i in range(N):
        for j in range(min(M,16384)):
            # print(sheet.cell(j+1,i+1).value)
            new_sheet.cell(i+1,j+1).value=sheet.cell(j+1,i+1).value

    new_wb.save(f'inverted{file}')

if __name__=="__main__":
    if len(sys.argv)>1:
        file=sys.argv[1]
    else:
        print('Give file that needs to inverted.')
        sys.exit()

    blankRows(file)