import openpyxl,sys,os

def txt2excel(folder):
    new_wb=openpyxl.Workbook()
    new_sheet=new_wb[new_wb.sheetnames[0]]
    for i,file in enumerate(os.listdir(folder)):
        with open(folder+'/'+file,'r') as f:
            lines=f.readlines()
        # give file name as column name
        new_sheet.cell(1,i+1).value=file.split('.')[0]
        for j,line in enumerate(lines):
            new_sheet.cell(j+2,i+1).value=line.strip()

    new_wb.save(f'txt2excel.xlsx')

if __name__=="__main__":
    if len(sys.argv)>1:
        folder=sys.argv[1]
    else:
        print('Give folder that contains files.')
        sys.exit()

    txt2excel(folder)