import openpyxl,sys
from openpyxl.utils import get_column_letter

def blankRows(N,M,file):
    wb=openpyxl.load_workbook(file)
    sheet=wb[wb.sheetnames[0]]

    new_wb=openpyxl.Workbook()
    new_sheet=new_wb[new_wb.sheetnames[0]]

    '''
    simple re-write approach
    '''
    for i,row in enumerate(sheet.rows):
        if i<N-1:
            for k,ele in enumerate(row):
                new_sheet[f'{get_column_letter(k+1)}{i+1}']=ele.value
        else:
            for k,ele in enumerate(row):
                new_sheet[f'{get_column_letter(k+1)}{i+1+M}']=ele.value

    new_wb.save(f'blankRows{file}')

if __name__=="__main__":
    if len(sys.argv)>3:
        N=sys.argv[1]
        M=sys.argv[2]
        file=sys.argv[3]
        if N.isnumeric() and M.isnumeric():
            N=int(N)
            M=int(M)
        else:
            print('Give valid values for N and M')
            sys.exit()
    else:
        print('Give N.')
        sys.exit()

    blankRows(N,M,file)